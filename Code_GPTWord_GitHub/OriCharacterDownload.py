from openpyxl import load_workbook
from utils import unique_test
from utils import read_txt, get_this_searchWord, unique, CheckFix_MultiSameWords
import os
from utils import get_filelist_frompath
import openpyxl

sep = os.sep

root_dataset_xlsx = r'D:\Work_Skyer\Lab\Doctor_ZYJ\GPTWords\data\root_dataset_xlsx'
TimeLoad = '20230402'
TimeSave = '20230402'
path_ori_xlsxLoad = root_dataset_xlsx + sep + 'dataset_xlsx_' + TimeLoad + '.xlsx'
path_new_xlsxSave = root_dataset_xlsx + sep + 'dataset_xlsx_' + TimeSave + '.xlsx'
path_ori_xlsxTemp = root_dataset_xlsx + sep + 'dataset_xlsx_Temp.xlsx'

root_download = r'D:\Work_Skyer\Lab\Doctor_ZYJ\GPTWords\data\DownloadDataRoot\20230104_MidTotal1'
path_RecordDownloadedSingleCharacter = r'D:\Work_Skyer\Lab\Doctor_ZYJ\GPTWords\data\RecordDownloadedSingleCharacter.txt'

do_Before_checkDataset = False
do_After_checkDataset = True

if do_Before_checkDataset:
    CheckFix_MultiSameWords(path_ori_xlsxLoad)

dataTxt_path_RecordDownloadedSingleCharacter = read_txt(path_RecordDownloadedSingleCharacter,encoding='ANSI')
ListTxt_path_RecordDownloadedSingleCharacter = list(dataTxt_path_RecordDownloadedSingleCharacter)
if not unique_test(ListTxt_path_RecordDownloadedSingleCharacter):
    print('path_RecordDownloadedSingleCharacter Wrong! Exist Multi-Same Words')
    unique_ListTxt_path_RecordDownloadedSingleCharacter = list(unique(ListTxt_path_RecordDownloadedSingleCharacter))
    file2 = open(path_RecordDownloadedSingleCharacter, "w")
    file2.write("".join(unique_ListTxt_path_RecordDownloadedSingleCharacter))
    ListTxt_path_RecordDownloadedSingleCharacter = unique_ListTxt_path_RecordDownloadedSingleCharacter
    file2.close()

file2 = open(path_RecordDownloadedSingleCharacter, "a",)

wb = load_workbook(path_ori_xlsxLoad)
sheets = wb.worksheets  # 获取当前所有的sheet
sheet0 = sheets[0]

# 新建工作簿：
wb_newSave = openpyxl.Workbook()
wb_newSave.save(path_new_xlsxSave)

max_row = sheet0.max_row
max_column = sheet0.max_column

# First Check
if sheet0.cell(1, 1).value == '词' and sheet0.cell(1, 2).value == '出现次数' and sheet0.cell(1, 3).value == '频率‰（千分之一）':
    pass
else:
    print('标题错误,进行修改')
    sheet0['A1'] = '词'
    sheet0['B1'] = '出现次数'
    sheet0['C1'] = '频率‰（千分之一）'

path_list = get_filelist_frompath(root_download, 'txt')

for index_path_DownloadTxt in range(len(path_list)):
    path_oneTxt_download = path_list[index_path_DownloadTxt]
    print(str(index_path_DownloadTxt + 1) + '/' + str(len(path_list)), 'Path=', path_oneTxt_download)

    one_StrTxt_download = read_txt(path_oneTxt_download, encoding='utf-8')
    one_ListTxt_download = one_StrTxt_download.split('\n')

    len_list_oneCharacter = len(one_ListTxt_download)
    if len_list_oneCharacter <= 6:
        print('txt_data_wrong!!!', path_oneTxt_download)

    this_Character = get_this_searchWord(the_ListTxt_download=one_ListTxt_download)
    if this_Character not in ListTxt_path_RecordDownloadedSingleCharacter:
        ListTxt_path_RecordDownloadedSingleCharacter.append(this_Character)
        file2.write(this_Character)
        for index_row in range(6, len_list_oneCharacter - 1):
            oneList_Word = list(one_ListTxt_download[index_row].split('\t\t'))
            this_word = oneList_Word[1]
            this_ExistNum = oneList_Word[2]
            this_ExistProbability = oneList_Word[3]

            sheet0.append([this_word, this_ExistNum, this_ExistProbability])
            # print('Add word ', this_word, ' into dataset_xlsx_' + TimeSave + '.xlsx')

file2.close()
wb.save(path_new_xlsxSave)

if do_After_checkDataset:
    CheckFix_MultiSameWords(path_new_xlsxSave)