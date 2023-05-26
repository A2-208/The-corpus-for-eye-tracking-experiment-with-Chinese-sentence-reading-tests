from utils import Test_Readability_OneChineseContent_Office_V1, xlsx_clearAll, wxwx
from openpyxl import load_workbook, Workbook
import os

Num_Multi_Words = 1
type_prop = 'n'
print('type_prop=', type_prop)

root_ChatGPT_document = r'D:\Work_Skyer\Lab\Doctor_ZYJ\GPTWords'

path_dataset_xlsxLoad = root_ChatGPT_document+'\data\Root_Dataset_xlsx\dataset_xlsx_20230426_Main.xlsx'
root_Temp_download = root_ChatGPT_document+'\data\DownloadDataRoot\Temp'

# path_preprocess_xlsx = root_ChatGPT_document+'\data\Experiment\LTXZ\Exp\LTXZ_Dataset_Case_60_' + type_prop + '_' + str(Num_Multi_Words) + '.xlsx'
# path_result_xlsx = root_ChatGPT_document+'\data\Result\LTXZ\LTXZ_Dataset_Case_60_' + type_prop + '_' + str(Num_Multi_Words) + '.xlsx'
path_preprocess_xlsx = root_ChatGPT_document+'\data\Experiment\ThreeBody\Exp\ThreeBody_Dataset_Case_60_' + type_prop + '_' + str(Num_Multi_Words) + '.xlsx'
path_result_xlsx = root_ChatGPT_document+'\data\Result\ThreeBody\ThreeBody_Dataset_Case_60_' + type_prop + '_' + str(Num_Multi_Words) + '.xlsx'

DatasetUpdateMode = 'Word'  # ='Word',or ='Character',用于Update的检索方式, Update的参数
AllScoreDivideByNum_Words_Or_Characters = 'Words'  # ='Words',or ='Characters',Words为分词后的词数，C为总字数


print_NoTExistInDataset_Words = False
do_Before_checkDataset = False
do_After_checkDataset = False

xlsx_result_ClearAll = True
if xlsx_result_ClearAll:
    xlsx_clearAll(path_result_xlsx)

wb = load_workbook(path_preprocess_xlsx)
sheets = wb.worksheets  # 获取当前所有的sheet
sheet0 = sheets[0]
max_row = sheet0.max_row

content_str_Original_list = []

content_str_Preprocessed_list_Ori_MultiWord = []
content_str_Preprocessed_list_New_MultiWord = []
Score_list_Ori_MultiWord = []
Score_list_New_MultiWord = []

start_row_xlsx = 2
for i in range(start_row_xlsx, max_row + 1):
    one_Sentence = sheet0.cell(i, 1).value  # (row_Index, column_Index)
    content_str_Original_list.append(one_Sentence)

for a in range(Num_Multi_Words):
    content_str_Preprocessed_list_Ori = []
    content_str_Preprocessed_list_New = []

    for i in range(start_row_xlsx, max_row + 1):
        one_Word_Ori = sheet0.cell(i, a * 4 + 2).value  # (row_Index, column_Index)
        one_Word_New = sheet0.cell(i, a * 4 + 3).value  # (row_Index, column_Index)

        content_str_Preprocessed_list_Ori.append(one_Word_Ori)
        content_str_Preprocessed_list_New.append(one_Word_New)

    Score_list_Ori = Test_Readability_OneChineseContent_Office_V1(
        content_str_Preprocessed_list=content_str_Preprocessed_list_Ori,
        path_xlsxLoad=path_dataset_xlsxLoad,
        AllScoreDivideByNum_Words_Or_Characters=AllScoreDivideByNum_Words_Or_Characters,  # ='Words',or ='Characters',Words为分词后的词数，C为总字数
        print_NoTExistInDataset_Words=print_NoTExistInDataset_Words,
    )

    Score_list_New = Test_Readability_OneChineseContent_Office_V1(
        content_str_Preprocessed_list=content_str_Preprocessed_list_New,
        path_xlsxLoad=path_dataset_xlsxLoad,
        AllScoreDivideByNum_Words_Or_Characters=AllScoreDivideByNum_Words_Or_Characters,  # ='Words',or ='Characters',Words为分词后的词数，C为总字数
        print_NoTExistInDataset_Words=print_NoTExistInDataset_Words,
    )

    if len(Score_list_Ori) != len(Score_list_Ori):
        print('Num Ori and New has Wrong!!!')

    content_str_Preprocessed_list_Ori_MultiWord.append(content_str_Preprocessed_list_Ori)
    content_str_Preprocessed_list_New_MultiWord.append(content_str_Preprocessed_list_New)
    Score_list_Ori_MultiWord.append(Score_list_Ori)
    Score_list_New_MultiWord.append(Score_list_New)

if os.path.exists(path_result_xlsx) is not True:
    wb = Workbook()
    wb.create_sheet(index=0, title="ContentRecord")
    wb.save(path_result_xlsx)

wx_result = wxwx(path_result_xlsx)
wx_result.write_one_value("原句", 1, 1)
for i in range(Num_Multi_Words):
    wx_result.write_one_value("原词" + str(i), 1, 1 + i * 4 + 1)
    wx_result.write_one_value("新词" + str(i), 1, 1 + i * 4 + 2)
    wx_result.write_one_value("原词词频" + str(i), 1, 1 + i * 4 + 3)
    wx_result.write_one_value("新词词频" + str(i), 1, 1 + i * 4 + 4)

ALL_SCORE_ori_MultiWord = []
ALL_SCORE_new_MultiWord = []
for a in range(Num_Multi_Words):
    ALL_SCORE_ori_MultiWord.append(0.)
    ALL_SCORE_new_MultiWord.append(0.)

#     Score_list_Ori = Score_list_Ori_MultiWord[a]
#     Score_list_New = Score_list_New_MultiWord[a]
#     content_str_Preprocessed_list_Ori = content_str_Preprocessed_list_Ori_MultiWord[a]
#     content_str_Preprocessed_list_New = content_str_Preprocessed_list_New_MultiWord[a]
Num_case = len(Score_list_Ori_MultiWord[0])
for i in range(Num_Multi_Words):
    Num_c_ori = len(Score_list_Ori_MultiWord[i])
    Num_c_new = len(Score_list_New_MultiWord[i])
    if Num_c_ori != Num_case or Num_c_new != Num_case:
        print('Wrong!!! Score_list Num_case has Problem!!!')

# print(Score_list)
for one_index_content in range(Num_case):
    the_sentence = content_str_Original_list[one_index_content]
    row_word_unit_list = [the_sentence]
    for a in range(Num_Multi_Words):
        Score_list_Ori = Score_list_Ori_MultiWord[a]
        Score_list_New = Score_list_New_MultiWord[a]
        content_str_Preprocessed_list_Ori = content_str_Preprocessed_list_Ori_MultiWord[a]
        content_str_Preprocessed_list_New = content_str_Preprocessed_list_New_MultiWord[a]

        the_word_ori = content_str_Preprocessed_list_Ori[one_index_content]
        the_word_new = content_str_Preprocessed_list_New[one_index_content]

        the_score_ori = Score_list_Ori[one_index_content]
        the_score_new = Score_list_New[one_index_content]

        one_word_unit = [the_word_ori, the_word_new, the_score_ori, the_score_new]
        row_word_unit_list += one_word_unit

        ALL_SCORE_ori_MultiWord[a] += the_score_ori
        ALL_SCORE_new_MultiWord[a] += the_score_new

    wx_result.append_value_list(row_word_unit_list)

for a in range(Num_Multi_Words):
    ALL_SCORE_ori_MultiWord[a] = ALL_SCORE_ori_MultiWord[a] / Num_case
    ALL_SCORE_new_MultiWord[a] = ALL_SCORE_new_MultiWord[a] / Num_case

    print('ALL_SCORE_ori=', round(ALL_SCORE_ori_MultiWord[a], ndigits=7), '\t', 'WordChange(' + str(a) + ')', 'For content_str_list of :', Num_case,
          'Sentences')
    print('ALL_SCORE_new=', round(ALL_SCORE_new_MultiWord[a], ndigits=7), '\t', 'WordChange(' + str(a) + ')', 'For content_str_list of :', Num_case,
          ' Sentences')

wx_result.save_xlsx_cover(wx_result.path_xlsx)
