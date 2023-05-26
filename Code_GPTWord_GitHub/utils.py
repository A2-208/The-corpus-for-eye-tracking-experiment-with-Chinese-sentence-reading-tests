from openpyxl import load_workbook, Workbook
import os
import jieba
import random
import time

sep = os.sep


def xlsx_clearAll(path_x):
    x = Workbook()
    x.save(path_x)


def txt_clearAll(path_txt, encoding='utf-8'):
    txt = open(path_txt, 'w+', encoding=encoding)
    txt.write('')
    txt.close()


def get_one_WordsAndVn(one_WordSeries: str):
    d = one_WordSeries.split('/')
    d0 = d[0]
    d1 = d[1]

    return d0, d1


def make_vn(path_dataset_xlsx, path_txt):
    ww = wxwx(path_xlsx=path_dataset_xlsx)

    a = read_txt(path_txt, )
    b = a.split('\n')

    for i in range(0, len(b)):
        c = b[i]
        d = c.split('/')
        d0 = d[0]
        d1 = d[1]
        d10 = list(d1)
        d10A = ''.join(d10[0:-2])
        ww.write_one_value(value_cover='词性', row_index=1, column_str='D')
        print('Word=', d0, ' d10A=', d10A)
        if d0 == ww.get_one_value(row_index=i + 2, column_index=1, sheet_index=0):
            ww.write_one_value(value_cover=d10A, row_index=i + 2, column_str='D')

    ww.save_xlsx_cover(path_dataset_xlsx)


class OneWordClass:
    def __init__(self, word, probability_word, property_word):
        self.word = word
        self.probability_word = probability_word
        self.property_word = property_word


def get_sort_ByProbability_OfOneProperty(sentence_OriCut: str, ww_Source_Dataset, type_prop_list=None):
    if type_prop_list is None:
        type_prop_list = ['n', 'a', 'v']
    sentence_cut = sentence_OriCut.split('  ')[0:-1]
    sentence_cut = unique(sentence_cut)
    list_WordClass = []
    for one_WordSeries in sentence_cut:
        word, prop = get_one_WordsAndVn(one_WordSeries=one_WordSeries)
        if is_chinese(word):
            proba = ww_Source_Dataset.get_Probability_of_OneWord(word)

            # print('word=', word)
            # print('proba=', proba)
            # print('prop=', prop)

            if prop in type_prop_list:
                one_word_class = OneWordClass(word=word, probability_word=proba, property_word=prop)
                list_WordClass.append(one_word_class)

    # 进行从大到小的排序，list_WordClass操作完毕
    len_list_WordClass_Noun = len(list_WordClass)
    for a in range(0, len_list_WordClass_Noun):
        for b in range(a, len_list_WordClass_Noun):
            prob_highest = float(list_WordClass[a].probability_word)
            prob_new = float(list_WordClass[b].probability_word)
            if prob_new > prob_highest:
                list_WordClass[a], list_WordClass[b] = list_WordClass[b], list_WordClass[a]

    return list_WordClass


class wxwx:
    def __init__(self, path_xlsx):
        self.path_xlsx = path_xlsx
        if os.path.exists(self.path_xlsx) is not True:
            wb = Workbook()
            wb.save(self.path_xlsx)

        self.wb = load_workbook(self.path_xlsx)
        self.sheets = self.wb.worksheets  # 获取当前所有的sheet

    def get_one_value(self, row_index: int, column_index: int, sheet_index=0):
        return self.sheets[sheet_index].cell(row_index, column_index).value

    def write_one_value(self, value_cover: str, row_index: int, column_index: int, sheet_index=0):
        self.sheets[sheet_index].cell(row_index, column_index).value = value_cover

    def save_xlsx_cover(self, path_save: str):
        self.wb.save(path_save)

    def append_value_list(self, value_list: list, sheet_index=0):
        self.sheets[sheet_index].append(value_list)


class Source_Dataset:
    def __init__(self, path_xlsx_dataset):
        self.path_xlsx = path_xlsx_dataset

        self.wb = load_workbook(self.path_xlsx)
        self.sheets = self.wb.worksheets  # 获取当前所有的sheet
        self.sheet_0 = self.sheets[0]
        self.max_row = self.sheet_0.max_row

        # 收录content到list中
        self.ListExistWord = []
        self.ListExistWord_Probability = []
        # self.ListExistWord_Property = []
        for index_row in range(2, self.max_row + 1):
            one_word_exist = self.sheet_0.cell(index_row, 1).value  # (row_Index, column_Index)
            one_word_exist_Probability = self.sheet_0.cell(index_row, 3).value  # (row_Index, column_Index)
            # one_word_exist_Property = self.sheet_0.cell(index_row, 4).value  # (row_Index, column_Index)

            self.ListExistWord.append(one_word_exist)
            self.ListExistWord_Probability.append(one_word_exist_Probability)
            # self.ListExistWord_Property.append(one_word_exist_Property)

    def get_Probability_of_OneWord(self, word_one_search: str):
        if word_one_search in self.ListExistWord:
            this_index = self.ListExistWord.index(word_one_search)
            this_word_Probability = self.ListExistWord_Probability[this_index]
        else:
            this_word_Probability = 0.0
        return this_word_Probability

    # def get_Property_of_OneWord(self, word_one_search: str):
    #     this_index = self.ListExistWord.index(word_one_search)
    #     this_word_Property = self.ListExistWord_Property[this_index]
    #     return this_word_Property


def del_FilesOfOneFile(dir_path):
    if os.path.exists(dir_path):
        # os.walk会得到dir_path下各个后代文件夹和其中的文件的三元组列表，顺序自内而外排列，
        # 如 log下有111文件夹，111下有222文件夹：[('D:\\log\\111\\222', [], ['22.py']), ('D:\\log\\111', ['222'], ['11.py']), ('D:\\log', ['111'], ['00.py'])]
        for root, dirs, files in os.walk(dir_path, topdown=False):
            # print(root) # 各级文件夹绝对路径
            # print(dirs) # root下一级文件夹名称列表，如 ['文件夹1','文件夹2']
            # print(files)  # root下文件名列表，如 ['文件1','文件2']
            # 第一步：删除文件
            for name in files:
                os.remove(os.path.join(root, name))  # 删除文件


def Get_RecordedCharacterInDataset(path_xlsxLoad):
    wb = load_workbook(path_xlsxLoad)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet = sheets[1]
    max_row = sheet.max_row

    ListTxt_path_RecordDownloadedSingleCharacter = []
    for index_row in range(1, max_row + 1):
        one_word_exist = sheet.cell(index_row, 1).value  # (row_Index, column_Index)
        ListTxt_path_RecordDownloadedSingleCharacter.append(one_word_exist)
    ListTxt_path_RecordDownloadedSingleCharacter = unique_ChineseCharacters(
        ListTxt_path_RecordDownloadedSingleCharacter)
    return ListTxt_path_RecordDownloadedSingleCharacter


def SentenceToWords_list(sentence_str, cut_all=False):
    cut_res_list = list(jieba.cut(sentence_str, cut_all=cut_all))
    return cut_res_list





def get_filelist_frompath(filepath, expname):
    """
	读取文件夹中带有固定扩展名的文件
	:param filepath:
	:param expname: 扩展名，如'h5','PNG'
	:return: 文件路径list
	"""
    file_name = os.listdir(filepath)
    file_List = []
    for file in file_name:
        if file.endswith('.' + expname):
            file_List.append(os.path.join(filepath, file))
    return file_List


def get_this_searchWord(the_ListTxt_download):
    this_searchWord = the_ListTxt_download[3].split(' ')[-1]
    return this_searchWord


def is_all_chinese(strs):
    for _char in strs:
        if not '\u4e00' <= _char <= '\u9fa5':
            return False
    return True


def is_chinese(str):
    if str is None:
        return False
    if not '\u4e00' <= str <= '\u9fa5':
        return False
    return True


def unique(it):
    s = set()
    for x in it:
        if x not in s:
            s.add(x)
    return list(s)


def unique_ChineseCharacters(it):
    s = set()
    for x in it:
        if x not in s and is_chinese(x):
            s.add(x)
    return list(s)


def To_ChineseCharacters(it):
    s = []
    for x in it:
        if is_chinese(x):
            s.append(x)
    return s


def unique_test(it):
    s = set()
    for x in it:
        if x not in s:
            s.add(x)
    len_ori = len(it)
    len_s = len(s)
    if len_ori == len_s:
        return True
    else:
        return False


def read_txt(path_txt, encoding='utf-8'):  # or encoding='ANSI'
    # 方法一:读取每一行
    with open(path_txt, "r", encoding=encoding) as f:  # 打开文件
        data = f.read()  # 读取文件

    return data


def cover_txt(path_txt, encoding='utf-8'):
    # 方法一:读取每一行
    with open(path_txt, "w", encoding=encoding) as f:  # 打开文件
        data = f.read()  # 读取文件

    return data


def CheckFix_UniqueCharacter_Txt(path_txt, encoding='utf-8'):  # or encoding='ANSI'
    dataTxt_path_RecordDownloadedSingleCharacter = read_txt(path_txt, encoding=encoding)
    ListTxt_path_RecordDownloadedSingleCharacter = list(dataTxt_path_RecordDownloadedSingleCharacter)
    unique_ListTxt_path_RecordDownloadedSingleCharacter = unique(ListTxt_path_RecordDownloadedSingleCharacter)
    DownloadedSingleCharacter_txt = open(path_txt, "w")
    DownloadedSingleCharacter_txt.write("".join(unique_ListTxt_path_RecordDownloadedSingleCharacter))
    DownloadedSingleCharacter_txt.close()


def CheckFix_MultiSameWords(path_xlsxLoad):
    wb = load_workbook(path_xlsxLoad)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet0 = sheets[0]
    max_row = sheet0.max_row

    # First Check
    if sheet0.cell(1, 1).value == '词' and sheet0.cell(1, 2).value == '出现次数' and sheet0.cell(1, 3).value == '频率‰（千分之一）':
        pass
    else:
        print('标题错误,进行修改')
        sheet0['A1'] = '词'
        sheet0['B1'] = '出现次数'
        sheet0['C1'] = '频率‰（千分之一）'

    ListExistWord = []
    ListExistWord_Num = []
    ListExistWord_Proba = []

    ListExistWord_unique = []
    ListExistWord_Num_unique = []
    ListExistWord_Proba_unique = []

    for index_row in range(2, max_row + 1):
        one_word_exist = sheet0.cell(index_row, 1).value  # (row_Index, column_Index)
        one_Num_exist = sheet0.cell(index_row, 2).value  # (row_Index, column_Index)
        one_Proba_exist = sheet0.cell(index_row, 3).value  # (row_Index, column_Index)

        ListExistWord.append(one_word_exist)
        ListExistWord_Num.append(one_Num_exist)
        ListExistWord_Proba.append(one_Proba_exist)

        if one_word_exist not in ListExistWord_unique:
            ListExistWord_unique.append(one_word_exist)
            ListExistWord_Num_unique.append(one_Num_exist)
            ListExistWord_Proba_unique.append(one_Proba_exist)

    UniqueListExistWord = unique(ListExistWord)
    Num_UniqueListExistWord = len(UniqueListExistWord)

    if len(ListExistWord) != UniqueListExistWord:
        for index_row in range(2, max_row + 1):
            if index_row <= Num_UniqueListExistWord + 1:
                sheet0['A' + str(index_row)] = ListExistWord_unique[index_row - 2]
                sheet0['B' + str(index_row)] = ListExistWord_Num_unique[index_row - 2]
                sheet0['C' + str(index_row)] = ListExistWord_Proba_unique[index_row - 2]
            else:
                sheet0['A' + str(index_row)] = None
                sheet0['B' + str(index_row)] = None
                sheet0['C' + str(index_row)] = None

    wb.save(path_xlsxLoad)


def TransDownload_To_DatasetXlsx(path_xlsxLoad,
                                 root_download,
                                 do_Before_checkDataset=False,
                                 do_After_checkDataset=False, ):
    if do_Before_checkDataset:
        CheckFix_MultiSameWords(path_xlsxLoad)

    wb = load_workbook(path_xlsxLoad)
    sheets = wb.worksheets  # 获取当前所有的sheet
    if len(sheets) < 2:
        wb.create_sheet(index=1, title="Character")
        sheets = wb.worksheets  # 获取当前所有的sheet

    # print('len(sheets)',len(sheets))
    sheet0 = sheets[0]

    sheet1 = sheets[1]
    max_row1 = sheet1.max_row
    ListTxt_path_RecordDownloadedSingleCharacter = []
    for index_row in range(1, max_row1 + 1):
        one_word_exist = sheet1.cell(index_row, 1).value  # (row_Index, column_Index)
        if one_word_exist is not None:
            ListTxt_path_RecordDownloadedSingleCharacter.append(one_word_exist)

    # First Check
    if sheet0.cell(1, 1).value == '词' and sheet0.cell(1, 2).value == '出现次数' and sheet0.cell(1, 3).value == '频率‰（千分之一）':
        pass
    else:
        print('标题错误,进行修改')
        sheet0['A1'] = '词'
        sheet0['B1'] = '出现次数'
        sheet0['C1'] = '频率‰（千分之一）'

    path_list = get_filelist_frompath(root_download, 'txt')

    for index_path_DownloadTxt in range(len(path_list)):
        path_oneTxt_download = path_list[index_path_DownloadTxt]
        print(str(index_path_DownloadTxt + 1) + '/' + str(len(path_list)), 'Path=', path_oneTxt_download)

        one_StrTxt_download = read_txt(path_oneTxt_download, encoding='utf-8')
        one_ListTxt_download = one_StrTxt_download.split('\n')

        len_list_oneCharacter = len(one_ListTxt_download)
        if len_list_oneCharacter <= 6:
            print('txt_data_wrong!!!', path_oneTxt_download)

        this_Character = get_this_searchWord(the_ListTxt_download=one_ListTxt_download)
        # print(this_Character)
        # print(ListTxt_path_RecordDownloadedSingleCharacter)

        sheet1.append([this_Character])
        ListTxt_path_RecordDownloadedSingleCharacter.append(this_Character)

        for index_row in range(6, len_list_oneCharacter - 1):
            oneList_Word = list(one_ListTxt_download[index_row].split('\t\t'))
            this_word = oneList_Word[1]
            this_ExistNum = oneList_Word[2]
            this_ExistProbability = oneList_Word[3]

            sheet0.append([this_word, this_ExistNum, this_ExistProbability])
            # print('Add word ', this_word, ' into dataset')

    # print(ListTxt_path_RecordDownloadedSingleCharacter)
    wb.save(path_xlsxLoad)

    if do_After_checkDataset:
        print('Checking Dataset! Don\'t Cut Program! As It May Damage The Dataset Xlsx!!!')
        CheckFix_MultiSameWords(path_xlsxLoad)
        CheckFix_MultiSameCharacters(path_xlsxLoad)


def CheckFix_MultiSameCharacters(path_xlsxLoad):
    wb = load_workbook(path_xlsxLoad)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet = sheets[1]
    max_row = sheet.max_row
    # print('max_row',max_row)

    ListTxt_path_RecordDownloadedSingleCharacter = []
    for index_row in range(1, max_row + 1):
        one_word_exist = sheet.cell(index_row, 1).value  # (row_Index, column_Index)
        ListTxt_path_RecordDownloadedSingleCharacter.append(one_word_exist)

    Unique_ListTxt_path_RecordDownloadedSingleCharacter = unique_ChineseCharacters(
        ListTxt_path_RecordDownloadedSingleCharacter)
    len_Unique_ListTxt_path_RecordDownloadedSingleCharacter = len(Unique_ListTxt_path_RecordDownloadedSingleCharacter)
    if len(ListTxt_path_RecordDownloadedSingleCharacter) != len_Unique_ListTxt_path_RecordDownloadedSingleCharacter:
        for index_row in range(1, max_row + 1):
            if index_row <= len_Unique_ListTxt_path_RecordDownloadedSingleCharacter:
                sheet['A' + str(index_row)] = Unique_ListTxt_path_RecordDownloadedSingleCharacter[index_row - 1]
            else:
                sheet['A' + str(index_row)] = None

    wb.save(path_xlsxLoad)


def Test_sentence_ReadabilityScore_from_DatasetXlsx(test_content_str, the_path_Dataset_xlsx,
                                                    AllScoreDivideByNum_Words_Or_Characters='Words',
                                                    # ='Words',or ='Characters',Words为分词后的词数，C为总字数
                                                    print_NoTExistInDataset_Words=False,
                                                    cut_all=False):  # cut_all需==False，否则会有问题
    wb = load_workbook(the_path_Dataset_xlsx)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet0 = sheets[0]
    max_row = sheet0.max_row

    ListExistWord = []
    ListExistWord_RowIndexXlsx = []
    for index_row in range(2, max_row + 1):
        one_word_exist = sheet0.cell(index_row, 1).value  # (row_Index, column_Index)
        ListExistWord.append(one_word_exist)
        ListExistWord_RowIndexXlsx.append(index_row)

    test_words_list = SentenceToWords_list(sentence_str=test_content_str, cut_all=cut_all)
    test_words_list = To_ChineseCharacters(test_words_list)

    ListWords_Unique = unique(test_words_list)
    Len_Words_Unique = len(ListWords_Unique)
    ListWords_Num = [0 for i in range(Len_Words_Unique)]

    for one_word in test_words_list:
        the_index = ListWords_Unique.index(one_word)
        ListWords_Num[the_index] += 1

    NoTExistInDataset_Words_list = []
    ListWords_Probability = []
    for one_unique_words in ListWords_Unique:
        if one_unique_words in ListExistWord:
            this_index = ListExistWord.index(one_unique_words)
            this_row_in_xlsx = ListExistWord_RowIndexXlsx[this_index]
            # this_word_row_Num = sheet0.cell(this_row_in_xlsx, 2).value  # (row_Index, column_Index)
            this_word_row_Probability = sheet0.cell(this_row_in_xlsx, 3).value  # (row_Index, column_Index)
        else:
            NoTExistInDataset_Words_list.append(one_unique_words)
            this_word_row_Probability = 0.  # (row_Index, column_Index)

        ListWords_Probability.append(this_word_row_Probability)

    All_Score = 0.
    if AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        # 若计算character，则将每个Words的分数乘以“该Words所含的Character（汉字）的个数”，例如“电脑” 需x2， “过山车”需x3
        for one_unique_words_index in range(Len_Words_Unique):
            the_oneWord = ListWords_Unique[one_unique_words_index]
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]
            Num_CharactersOfOneWord = len(list(the_oneWord))

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord * Num_CharactersOfOneWord)
    else:
        for one_unique_words_index in range(Len_Words_Unique):
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord)

    if print_NoTExistInDataset_Words:
        num_NoTExistInDataset_Words_list = len(NoTExistInDataset_Words_list)
        if num_NoTExistInDataset_Words_list > 0:
            print(num_NoTExistInDataset_Words_list, ' Words not in Dataset,', 'Dataset have total ', len(ListExistWord),
                  ' Words', '以下词均不在词典统计中，计出现频率为0.0，（单位为千分之一）', NoTExistInDataset_Words_list)

    if AllScoreDivideByNum_Words_Or_Characters == 'Words':
        len_all = len(test_words_list)
        All_Score = All_Score / len_all
    elif AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        len_all = len(list(test_content_str))
        All_Score = All_Score / len_all
    else:
        print(
            'Choose a way to calculate the Score with correct Value (=\'Words\',or =\'Characters\',Words为分词后的词数，C为总字数)')
        print('Result Ori Total All Score Now')

    return All_Score



def random_wait_time(shortest_second=1.0, range_sec=None):
    if range_sec is None:
        range_sec = [0, 3]
    r_second = random.uniform(range_sec[0], range_sec[1])
    time.sleep(shortest_second + r_second)
    # print('sleep for ', shortest_second + r_second, ' second')


def Test_Readability_OneChineseContent(
        content_str_list,
        path_xlsxLoad,
        AllScoreDivideByNum_Words_Or_Characters='Words',
        print_NoTExistInDataset_Words=True,
):

    Score_out_list = []
    for one_content_str in content_str_list:
        Score = Test_sentence_ReadabilityScore_from_DatasetXlsx(test_content_str=one_content_str,
                                                                the_path_Dataset_xlsx=path_xlsxLoad,
                                                                AllScoreDivideByNum_Words_Or_Characters=AllScoreDivideByNum_Words_Or_Characters,
                                                                print_NoTExistInDataset_Words=print_NoTExistInDataset_Words)
        Score_out_list.append(Score)
    return Score_out_list


def CheckDownloadUpdate_Content_Office(
        content_str,  # 文本
        path_xlsxSave,
        root_Temp_download,  # 空文件夹用于暂时保存文件

        do_Before_checkDataset=False,
        do_After_checkDataset=True,
):
    print('Checking NoTExistInDataset')
    list_NoTExistInDataset = []

    wb = load_workbook(path_xlsxSave)
    sheets = wb.worksheets  # 获取当前所有的sheet

    sheet1 = sheets[1]
    max_row1 = sheet1.max_row

    ListExistCharacter = []
    for index_row in range(1, max_row1 + 1):
        one_character_exist = sheet1.cell(index_row, 1).value  # (row_Index, column_Index)
        ListExistCharacter.append(one_character_exist)
    list_UniqueCharacter = unique(To_ChineseCharacters(content_str))
    for one_UniqueCharacter in list_UniqueCharacter:
        if one_UniqueCharacter not in ListExistCharacter:
            if is_chinese(one_UniqueCharacter):
                list_NoTExistInDataset.append(one_UniqueCharacter)

    wb.close()

    if len(list_NoTExistInDataset) > 0:
        print('Dataset of \n' + path_xlsxSave + '\nlack Characters:', list_NoTExistInDataset)
        print('Downloading')

        del_FilesOfOneFile(dir_path=root_Temp_download)
        TransDownload_To_DatasetXlsx(path_xlsxLoad=path_xlsxSave,
                                     root_download=root_Temp_download,
                                     do_Before_checkDataset=do_Before_checkDataset,
                                     do_After_checkDataset=do_After_checkDataset,
                                     )
        print('Update Success')


def Test_sentence_ReadabilityScore_from_DatasetXlsx_Office(test_words_list, the_path_Dataset_xlsx,
                                                           AllScoreDivideByNum_Words_Or_Characters='Words',
                                                           # ='Words',or ='Characters',Words为分词后的词数，C为总字数
                                                           print_NoTExistInDataset_Words=False):  # cut_all需==False，否则会有问题
    wb = load_workbook(the_path_Dataset_xlsx)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet0 = sheets[0]
    max_row = sheet0.max_row

    # 收录content到list中
    ListExistWord = []
    ListExistWord_RowIndexXlsx = []
    for index_row in range(2, max_row + 1):
        one_word_exist = sheet0.cell(index_row, 1).value  # (row_Index, column_Index)
        ListExistWord.append(one_word_exist)
        ListExistWord_RowIndexXlsx.append(index_row)

    test_words_list = To_ChineseCharacters(test_words_list)

    ListWords_Unique = unique(test_words_list)
    Len_Words_Unique = len(ListWords_Unique)
    ListWords_Num = [0 for i in range(Len_Words_Unique)]

    for one_word in test_words_list:
        the_index = ListWords_Unique.index(one_word)
        ListWords_Num[the_index] += 1

    NoTExistInDataset_Words_list = []
    ListWords_Probability = []
    for one_unique_words in ListWords_Unique:
        if one_unique_words in ListExistWord:
            this_index = ListExistWord.index(one_unique_words)
            this_row_in_xlsx = ListExistWord_RowIndexXlsx[this_index]
            # this_word_row_Num = sheet0.cell(this_row_in_xlsx, 2).value  # (row_Index, column_Index)
            this_word_row_Probability = sheet0.cell(this_row_in_xlsx, 3).value  # (row_Index, column_Index)
        else:
            NoTExistInDataset_Words_list.append(one_unique_words)
            this_word_row_Probability = 0.  # (row_Index, column_Index)

        ListWords_Probability.append(this_word_row_Probability)

    All_Score = 0.
    if AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        # 若计算character，则将每个Words的分数乘以“该Words所含的Character（汉字）的个数”，例如“电脑” 需x2， “过山车”需x3
        for one_unique_words_index in range(Len_Words_Unique):
            the_oneWord = ListWords_Unique[one_unique_words_index]
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]
            Num_CharactersOfOneWord = len(list(the_oneWord))

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord * Num_CharactersOfOneWord)
    else:
        for one_unique_words_index in range(Len_Words_Unique):
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord)

    if print_NoTExistInDataset_Words:
        num_NoTExistInDataset_Words_list = len(NoTExistInDataset_Words_list)
        if num_NoTExistInDataset_Words_list > 0:
            print('\n', num_NoTExistInDataset_Words_list, ' Words not in Dataset,', 'Dataset have total ',
                  len(ListExistWord),
                  ' Words', '以下词均不在词典统计中，计出现频率为0.0，（单位为千分之一）', NoTExistInDataset_Words_list)

    if AllScoreDivideByNum_Words_Or_Characters == 'Words':
        len_all = len(test_words_list)
        All_Score = All_Score / len_all
    elif AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        len_all = len(list(''.join(test_words_list)))
        All_Score = All_Score / len_all
    else:
        print(
            'Choose a way to calculate the Score with correct Value (=\'Words\',or =\'Characters\',Words为分词后的词数，C为总字数)')
        print('Result Ori Total All Score Now')

    return All_Score


def Test_sentence_ReadabilityScore_from_DatasetXlsx_Office_V1(test_words_list, the_sheet0_Dataset_xlsx,
                                                              AllScoreDivideByNum_Words_Or_Characters='Words',
                                                              # ='Words',or ='Characters',Words为分词后的词数，C为总字数
                                                              print_NoTExistInDataset_Words=False):  # cut_all需==False，否则会有问题
    # wb = load_workbook(the_path_Dataset_xlsx)
    # sheets = wb.worksheets  # 获取当前所有的sheet
    # sheet0 = sheets[0]

    sheet0 = the_sheet0_Dataset_xlsx
    max_row = sheet0.max_row

    # 收录content到list中
    ListExistWord = []
    ListExistWord_RowIndexXlsx = []
    for index_row in range(2, max_row + 1):
        one_word_exist = sheet0.cell(index_row, 1).value  # (row_Index, column_Index)
        ListExistWord.append(one_word_exist)
        ListExistWord_RowIndexXlsx.append(index_row)

    test_words_list = To_ChineseCharacters(test_words_list)

    ListWords_Unique = unique(test_words_list)
    Len_Words_Unique = len(ListWords_Unique)
    ListWords_Num = [0 for i in range(Len_Words_Unique)]

    for one_word in test_words_list:
        the_index = ListWords_Unique.index(one_word)
        ListWords_Num[the_index] += 1

    NoTExistInDataset_Words_list = []
    ListWords_Probability = []
    for one_unique_words in ListWords_Unique:
        if one_unique_words in ListExistWord:
            this_index = ListExistWord.index(one_unique_words)
            this_row_in_xlsx = ListExistWord_RowIndexXlsx[this_index]
            # this_word_row_Num = sheet0.cell(this_row_in_xlsx, 2).value  # (row_Index, column_Index)
            this_word_row_Probability = sheet0.cell(this_row_in_xlsx, 3).value  # (row_Index, column_Index)
        else:
            NoTExistInDataset_Words_list.append(one_unique_words)
            this_word_row_Probability = 0.  # (row_Index, column_Index)

        ListWords_Probability.append(this_word_row_Probability)

    All_Score = 0.
    if AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        # 若计算character，则将每个Words的分数乘以“该Words所含的Character（汉字）的个数”，例如“电脑” 需x2， “过山车”需x3
        for one_unique_words_index in range(Len_Words_Unique):
            the_oneWord = ListWords_Unique[one_unique_words_index]
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]
            Num_CharactersOfOneWord = len(list(the_oneWord))

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord * Num_CharactersOfOneWord)
    else:
        for one_unique_words_index in range(Len_Words_Unique):
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord)

    if print_NoTExistInDataset_Words:
        num_NoTExistInDataset_Words_list = len(NoTExistInDataset_Words_list)
        if num_NoTExistInDataset_Words_list > 0:
            print('\n', num_NoTExistInDataset_Words_list, ' Words not in Dataset,', 'Dataset have total ',
                  len(ListExistWord),
                  ' Words', '以下词均不在词典统计中，计出现频率为0.0，（单位为千分之一）', NoTExistInDataset_Words_list)

    if AllScoreDivideByNum_Words_Or_Characters == 'Words':
        len_all = len(test_words_list)
        All_Score = All_Score / len_all
    elif AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        len_all = len(list(''.join(test_words_list)))
        All_Score = All_Score / len_all
    else:
        print(
            'Choose a way to calculate the Score with correct Value (=\'Words\',or =\'Characters\',Words为分词后的词数，C为总字数)')
        print('Result Ori Total All Score Now')

    return All_Score


def Test_sentence_ReadabilityScore_from_DatasetXlsx_Office_Faster(test_words_list, the_path_Dataset_xlsx,
                                                                  AllScoreDivideByNum_Words_Or_Characters='Words',
                                                                  # ='Words',or ='Characters',Words为分词后的词数，C为总字数
                                                                  print_NoTExistInDataset_Words=False):  # cut_all需==False，否则会有问题
    wb = load_workbook(the_path_Dataset_xlsx)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet0 = sheets[0]
    max_row = sheet0.max_row

    # 收录content到list中
    ListExistWord = []
    ListExistWord_Probability = []
    for index_row in range(2, max_row + 1):
        one_word_exist = sheet0.cell(index_row, 1).value  # (row_Index, column_Index)
        one_word_exist_Probability = sheet0.cell(index_row, 3).value  # (row_Index, column_Index)

        ListExistWord.append(one_word_exist)
        ListExistWord_Probability.append(one_word_exist_Probability)

    test_words_list = To_ChineseCharacters(test_words_list)

    ListWords_Unique = unique(test_words_list)
    Len_Words_Unique = len(ListWords_Unique)
    ListWords_Num = [0 for i in range(Len_Words_Unique)]

    for one_word in test_words_list:
        the_index = ListWords_Unique.index(one_word)
        ListWords_Num[the_index] += 1

    NoTExistInDataset_Words_list = []
    ListWords_Probability = []
    for one_unique_words in ListWords_Unique:
        if one_unique_words in ListExistWord:
            this_index = ListExistWord.index(one_unique_words)
            # this_row_in_xlsx = ListExistWord_RowIndexXlsx[this_index]
            this_word_row_Probability = ListExistWord_Probability[this_index]
        else:
            NoTExistInDataset_Words_list.append(one_unique_words)
            this_word_row_Probability = 0.  # (row_Index, column_Index)

        ListWords_Probability.append(this_word_row_Probability)

    All_Score = 0.
    if AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        # 若计算character，则将每个Words的分数乘以“该Words所含的Character（汉字）的个数”，例如“电脑” 需x2， “过山车”需x3
        for one_unique_words_index in range(Len_Words_Unique):
            the_oneWord = ListWords_Unique[one_unique_words_index]
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]
            Num_CharactersOfOneWord = len(list(the_oneWord))

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord * Num_CharactersOfOneWord)
    else:
        for one_unique_words_index in range(Len_Words_Unique):
            the_Pro_of_oneWord = ListWords_Probability[one_unique_words_index]
            the_ShowTimeOfThisSentence_of_oneWord = ListWords_Num[one_unique_words_index]

            All_Score += (float(the_Pro_of_oneWord) * the_ShowTimeOfThisSentence_of_oneWord)

    if print_NoTExistInDataset_Words:
        num_NoTExistInDataset_Words_list = len(NoTExistInDataset_Words_list)
        if num_NoTExistInDataset_Words_list > 0:
            print('\n', num_NoTExistInDataset_Words_list, ' Words not in Dataset,', 'Dataset have total ',
                  len(ListExistWord),
                  ' Words', '以下词均不在词典统计中，计出现频率为0.0，（单位为千分之一）', NoTExistInDataset_Words_list)

    if AllScoreDivideByNum_Words_Or_Characters == 'Words':
        len_all = len(test_words_list)
        All_Score = All_Score / len_all
    elif AllScoreDivideByNum_Words_Or_Characters == 'Characters':
        len_all = len(list(''.join(test_words_list)))
        All_Score = All_Score / len_all
    else:
        print(
            'Choose a way to calculate the Score with correct Value (=\'Words\',or =\'Characters\',Words为分词后的词数，C为总字数)')
        print('Result Ori Total All Score Now')

    return All_Score


def Test_Readability_OneChineseContent_Office_V1(
        content_str_Preprocessed_list,
        path_xlsxLoad,
        AllScoreDivideByNum_Words_Or_Characters='Words',
        print_NoTExistInDataset_Words=True,
):

    Score_out_list = []
    LenSentence = len(content_str_Preprocessed_list)

    # ww = Source_Dataset(path_xlsx_dataset=path_xlsxLoad)
    wb = load_workbook(path_xlsxLoad)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet0 = sheets[0]
    for one_index in range(LenSentence):
        one_content_Preprocessed_str = content_str_Preprocessed_list[one_index]
        test_words_list = To_ChineseCharacters(one_content_Preprocessed_str.split(' '))
        Score = Test_sentence_ReadabilityScore_from_DatasetXlsx_Office_V1(test_words_list=test_words_list,
                                                                          the_sheet0_Dataset_xlsx=sheet0,
                                                                          AllScoreDivideByNum_Words_Or_Characters=AllScoreDivideByNum_Words_Or_Characters,
                                                                          print_NoTExistInDataset_Words=print_NoTExistInDataset_Words)
        Score_out_list.append(Score)
        print(str(one_index + 1) + '/' + str(LenSentence), 'Score=' + str(Score), ' Has been tested. Sentence:',
              one_content_Preprocessed_str)
    return Score_out_list
